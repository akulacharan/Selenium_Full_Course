from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl


# Excell Structure  File-->Workbook-->Sheet-->Rows-->Cells
file = r"C:\Users\charanteja\PycharmProjects\SeleniumUpdatedCourse\Excel\SELENIUMLEARN.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook['Sheet1']
rows = sheet.max_row
cols = sheet.max_column

# Tip : You have to delete all the empty cells in the excell sheet first


for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end="       ")
    print()
