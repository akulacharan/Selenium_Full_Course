import openpyxl


''' To write same data in all the cells
file = r"C:\\Users\charanteja\PycharmProjects\SeleniumUpdatedCourse\Excel\EmptyExcelSheet.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active   # Gets active sheet in the Excel if we have only one sheet.
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(r,c).value="Cherry"
workbook.save(file)    # Save the file after entering the data
'''

# To add multiple data
file2 = r"C:\Users\charanteja\PycharmProjects\SeleniumUpdatedCourse\Excel\MultipleData.xlsx"
workbook = openpyxl.load_workbook(file2)
sheet = workbook.active   # Gets active sheet in the Excel if we have only one sheet.

sheet.cell(1,1).value=123
sheet.cell(1,2).value='charan'
sheet.cell(1,3).value='Engineer'

sheet.cell(2,1).value=456
sheet.cell(2,2).value='Arun'
sheet.cell(2,3).value='Tester'


sheet.cell(3,1).value=789
sheet.cell(3,2).value='Varun'
sheet.cell(3,3).value='Developer'

workbook.save(file2)